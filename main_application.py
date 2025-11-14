"""
메인 애플리케이션 클래스
"""
import sys
import threading
import time
from datetime import datetime
from typing import Dict, Any, Optional
import logging

from PyQt5.QtWidgets import (
    QApplication, QWidget, QDialog, QVBoxLayout, QHBoxLayout, 
    QLabel, QPushButton, QMessageBox, QProgressDialog
)
from PyQt5.QtCore import Qt, QSettings, QThread, pyqtSignal
from PyQt5.QtGui import QIcon

from config import (
    DARK_THEME_STYLE, DIR_PRESET, BUILD_NAME_FILE, FIX_VERSION_FILE, 
    SETTINGS_FILE, ensure_directories, AIConfig, EXCEL_EXPORT_FILE, APP_SETTINGS_FILE, DIR_RESULT
)
from utils import (
    FileManager, TextProcessor, PresetManager, ValidationHelper,
    OptionsManager, setup_logging
)
from gui_widgets import create_main_form, SettingsDialog
from jira_automation import create_issue, JiraAutomation
from ai_assistant import get_ai_assistant, AIAssistant

logger = logging.getLogger(__name__)


class AIGenerationThread(QThread):
    """AI 생성을 백그라운드에서 실행하는 스레드"""
    finished = pyqtSignal(dict)  # 생성 완료 시 결과 전달
    error = pyqtSignal(str)  # 에러 발생 시 메시지 전달
    
    def __init__(self, summary: str, preset_dir: str):
        super().__init__()
        self.summary = summary
        self.preset_dir = preset_dir
    
    def run(self):
        """스레드 실행"""
        try:
            # AI 어시스턴트 가져오기
            ai_assistant = get_ai_assistant(
                preset_dir=self.preset_dir,
                model_name=AIConfig.DEFAULT_MODEL
            )
            
            # AI로 버그 세부정보 생성
            result = ai_assistant.generate_bug_details(self.summary)
            
            if result:
                self.finished.emit(result)
            else:
                self.error.emit("AI가 유효한 응답을 생성하지 못했습니다.")
                
        except Exception as e:
            logger.error(f"AI 생성 중 오류: {e}", exc_info=True)
            self.error.emit(f"AI 생성 실패: {str(e)}")


class ExcelBatchThread(QThread):
    """엑셀 일괄 실행을 백그라운드에서 실행하는 스레드"""
    progress_update = pyqtSignal(int, str)  # 진행 상황 업데이트 (값, 메시지)
    issue_created = pyqtSignal(int, int, str)  # 이슈 생성 완료 (현재, 전체, 제목)
    error_occurred = pyqtSignal(int, str)  # 에러 발생 (행 번호, 에러 메시지)
    finished = pyqtSignal(int, list)  # 완료 (성공 개수, 실패 목록)
    
    def __init__(self, excel_path: str):
        super().__init__()
        self.excel_path = excel_path
        self.is_cancelled = False
    
    def cancel(self):
        """실행 취소"""
        self.is_cancelled = True
        logger.info("사용자가 일괄 실행 취소를 요청했습니다.")
    
    def run(self):
        """스레드 실행"""
        try:
            import openpyxl
            
            # 엑셀 파일 읽기
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            
            # 헤더 확인
            headers = [cell.value for cell in ws[1]]
            
            # JIRA Automation 인스턴스 생성
            jira_automation = JiraAutomation()
            jira_automation.start_driver()
            
            # 각 행을 순차적으로 처리
            success_count = 0
            failed_rows = []
            
            for row_idx in range(2, ws.max_row + 1):
                # 취소 확인
                if self.is_cancelled:
                    logger.info("일괄 실행이 취소되었습니다.")
                    break
                
                try:
                    # 행 데이터 읽기
                    row_data = [cell.value for cell in ws[row_idx]]
                    issue_data = self._parse_excel_row(headers, row_data)
                    
                    # 진행 상황 업데이트
                    summary_preview = issue_data.get('summary', '')[:50]
                    progress_text = (
                        f"이슈 생성 중... ({row_idx - 1}/{ws.max_row - 1})\n"
                        f"제목: {summary_preview}{'...' if len(issue_data.get('summary', '')) > 50 else ''}"
                    )
                    self.progress_update.emit(row_idx - 2, progress_text)
                    
                    # 새 탭 생성 (첫 번째 이슈가 아닐 경우)
                    if row_idx > 2:
                        jira_automation.create_new_tab()
                        time.sleep(1)
                    
                    # 이슈 생성
                    jira_automation.create_issue(issue_data, pause_for_review=False)
                    
                    logger.info(f"이슈 생성 완료 ({row_idx - 1}/{ws.max_row - 1}): {issue_data.get('summary', '')}")
                    success_count += 1
                    
                    # 이슈 생성 완료 시그널
                    self.issue_created.emit(row_idx - 1, ws.max_row - 1, issue_data.get('summary', ''))
                    
                    # 다음 이슈로 진행하기 전 짧은 대기
                    time.sleep(1.5)
                    
                except Exception as e:
                    logger.error(f"이슈 생성 실패 (행 {row_idx}): {e}", exc_info=True)
                    failed_rows.append((row_idx, str(e)))
                    
                    # 에러 발생 시그널
                    self.error_occurred.emit(row_idx, str(e))
            
            # 완료 시그널
            self.finished.emit(success_count, failed_rows)
            
            # 드라이버는 종료하지 않음 (사용자가 브라우저에서 확인 가능)
            # jira_automation.close()
            
        except Exception as e:
            logger.error(f"엑셀 일괄 실행 스레드 오류: {e}", exc_info=True)
            self.finished.emit(0, [(0, str(e))])
    
    def _parse_excel_row(self, headers: list, row_data: list) -> dict:
        """엑셀 행 데이터를 JIRA 이슈 데이터로 변환"""
        # 헤더와 데이터를 매핑
        data_dict = {}
        for i, header in enumerate(headers):
            if i < len(row_data):
                value = row_data[i]
                data_dict[header] = value if value is not None else ''
        
        # JIRA 필드명으로 변환
        issue_data = {
            'summary': str(data_dict.get('Summary', '')),
            'team': str(data_dict.get('Team', '')),
            'linkedIssues': str(data_dict.get('Linked Issues', '')),
            'issue': str(data_dict.get('Issue', '')),
            'parent': str(data_dict.get('Parent', '')),
            'reviewer': str(data_dict.get('Reviewer', '')),
            'branch': str(data_dict.get('Branch', '')),
            'build': str(data_dict.get('Build', '')),
            'fixversion': str(data_dict.get('Fix Version', '')),
            'component': str(data_dict.get('Component', '')),
            'label': str(data_dict.get('Label', '')),
            'priority': str(data_dict.get('Priority', '')),
            'severity': str(data_dict.get('Severity', '')),
            'prevalence': str(data_dict.get('Prevalence', '')),
            'repro_rate': str(data_dict.get('Repro Rate', '')),
            'steps': str(data_dict.get('Steps', '')),
            'description': str(data_dict.get('Description', ''))
        }
        
        return issue_data


class BugReportApp(QWidget):
    """버그 리포트 메인 애플리케이션 클래스"""
    
    def __init__(self):
        super().__init__()
        
        # 초기화
        setup_logging()
        ensure_directories()
        
        # 매니저들 초기화
        self.file_manager = FileManager()
        self.text_processor = TextProcessor()
        self.preset_manager = PresetManager(DIR_PRESET)
        self.options_manager = OptionsManager()
        
        # UI 요소들을 저장할 딕셔너리
        self.widgets = {}
        
        # AI 관련
        self.ai_thread = None
        self.ai_progress_dialog = None
        
        # 엑셀 일괄 실행 관련
        self.excel_batch_thread = None
        self.excel_progress_dialog = None
        
        # 앱 설정 (엑셀 추출 등)
        self.app_settings = self._load_app_settings()
        
        # UI 초기화
        self.init_ui()
        
        # 설정 로드
        self.load_settings()
    
    def init_ui(self):
        """UI를 초기화"""
        self.setWindowIcon(QIcon('jira_bug.ico'))
        self.setWindowTitle('Bug Report')
        self.setMinimumSize(500, 300)
        self.setStyleSheet(DARK_THEME_STYLE)
        
        # 폼 생성
        form_data = create_main_form(self, self.show_about_dialog, self.show_settings_dialog)
        
        # 위젯들 참조 저장
        self._store_widget_references(form_data)
        
        # 이벤트 연결
        self._connect_events()
        
        # 단축키 설정
        self._setup_shortcuts()
        
        # 레이아웃 설정
        self.setLayout(form_data['layout'])
        
        # 프리셋 새로고침
        self.refresh_presets()
        
        # 콤보 필드 옵션 초기화
        self._initialize_combo_field_options()
        
        self.show()
    
    def _store_widget_references(self, form_data: Dict[str, Any]):
        """위젯 참조들을 저장"""
        self.widgets.update({
            'preset_prefix': form_data['preset_widgets']['prefix_combo'],
            'preset_name': form_data['preset_widgets']['name_combo'],
            'preset_version': form_data['preset_widgets']['version_combo'],
            'preset_line': form_data['preset_widgets']['preset_line'],
            'preset_sort': form_data['preset_widgets']['sort_combo'],
            'generate_combo': form_data['action_widgets']['generate_combo'],
        })
        
        # 엑셀 일괄 실행 위젯들 저장
        if 'excel_batch' in form_data:
            self.excel_widgets = form_data['excel_batch']
        
        # 필드 위젯들
        self.other_fields = {}
        for field_name, field_data in form_data['field_widgets'].items():
            self.other_fields[field_name] = field_data['widget']
        
        # 새로운 콤보 필드들 (branch, build, fixversion, component)
        self.combo_field_widgets = {}
        for field_name, combo_data in form_data['combo_field_widgets'].items():
            self.other_fields[field_name] = combo_data['widget']  # 기존 방식과 호환
            self.combo_field_widgets[field_name] = combo_data['combo_field_widget']
            
            # 옵션 매니저 설정
            combo_data['combo_field_widget'].set_options_manager(self.options_manager)
            
        # summary 필드의 sub_label 처리
        if 'summary' in form_data['field_widgets']:
            field_widget = form_data['field_widgets']['summary']['field_widget']
            if hasattr(field_widget, 'sub_label'):
                self.sub_label = field_widget.sub_label
        
        # 콤보박스들
        for combo_name, combo_data in form_data['combo_widgets'].items():
            self.widgets[combo_name] = combo_data['widget']
        
        # 텍스트 필드들
        for text_name, text_data in form_data['text_widgets'].items():
            self.widgets[text_name] = text_data['widget']
        
        # 체크박스 (label 필드)
        if 'label' in form_data['field_widgets']:
            field_widget = form_data['field_widgets']['label']['field_widget']
            for btn_type, btn, _ in field_widget.extra_buttons:
                if btn_type == 'checkbox':
                    self.include_main_label_check_box = btn
                    break
        
        # 버튼들
        self.preset_buttons = form_data['preset_widgets']['buttons']
        self.action_buttons = form_data['action_widgets']['buttons']
        
        # 파일 관련 버튼들 저장
        self._store_file_buttons(form_data['field_widgets'])
    
    def _initialize_combo_field_options(self):
        """콤보 필드들의 옵션을 초기화"""
        for field_name, combo_field_widget in self.combo_field_widgets.items():
            combo_field_widget.refresh_options()
    
    def _store_file_buttons(self, field_widgets: Dict[str, Any]):
        """파일 관련 버튼들을 저장"""
        for field_name, field_data in field_widgets.items():
            field_widget = field_data['field_widget']
            for btn_type, btn, filename in field_widget.extra_buttons:
                if btn_type == 'load':
                    btn.clicked.connect(
                        lambda checked, f=filename, w=field_data['widget']: 
                        self._load_text_file(f, w)
                    )
                elif btn_type == 'save':
                    btn.clicked.connect(
                        lambda checked, f=filename, w=field_data['widget']:
                        self._save_text_file(f, w.text())
                    )
                elif btn_type == 'ai_generate':
                    # AI 생성 버튼 연결
                    btn.clicked.connect(self.generate_with_ai)
    
    def _connect_events(self):
        """이벤트들을 연결"""
        # 프리셋 버튼들
        self.preset_buttons['refresh'].clicked.connect(self.refresh_presets)
        self.preset_buttons['apply'].clicked.connect(self.apply_preset)
        self.preset_buttons['delete'].clicked.connect(self.delete_preset)
        self.preset_buttons['save_preset'].clicked.connect(self.save_preset)
        
        # 프리셋 콤보박스 변경
        self.widgets['preset_prefix'].currentIndexChanged.connect(lambda: self._on_prefix_changed())
        self.widgets['preset_name'].currentIndexChanged.connect(lambda: self._on_name_changed())
        self.widgets['preset_sort'].currentIndexChanged.connect(self.refresh_presets)
        
        # 콤보 필드 버튼들 연결
        for field_name, combo_field_widget in self.combo_field_widgets.items():
            combo_field_widget.add_button.clicked.connect(combo_field_widget.add_option)
            combo_field_widget.remove_button.clicked.connect(combo_field_widget.remove_option)
        
        # 액션 버튼들
        self.action_buttons['generate'].clicked.connect(self.generate_description)
        self.action_buttons['execute'].clicked.connect(self.execute)
        
        # 엑셀 일괄 실행 버튼들
        if hasattr(self, 'excel_widgets'):
            self.excel_widgets['open_file_btn'].clicked.connect(self.open_excel_file)
            self.excel_widgets['execute_btn'].clicked.connect(self.execute_excel_batch)
    
    def _setup_shortcuts(self):
        """단축키를 설정"""
        self.preset_buttons['refresh'].setShortcut('F5')
        self.preset_buttons['apply'].setShortcut('F6')
        self.action_buttons['execute'].setShortcut('F2')
    
    def _load_text_file(self, filename: str, target_widget):
        """텍스트 파일을 로드"""
        content = self.file_manager.load_text_file_all(filename)
        if content and target_widget:
            target_widget.setText(content)
    
    def _save_text_file(self, filename: str, content: str):
        """텍스트 파일을 저장"""
        self.file_manager.create_text_file(filename, content)
    
    def generate_description(self):
        """설명을 자동 생성"""
        main_text = self.other_fields['summary'].toPlainText()
        option = self.widgets['generate_combo'].currentText()
        build_text = self.other_fields['build'].currentText()
        
        description = self.text_processor.generate_description_template(
            main_text, option, build_text
        )
        
        self.widgets['description'].setText(description)
    
    def generate_with_ai(self):
        """AI로 버그 세부정보를 생성"""
        # AI 사용 가능 여부 확인
        if not AIAssistant.is_ollama_available():
            QMessageBox.warning(
                self, 
                "AI 기능 사용 불가",
                "ollama 패키지가 설치되지 않았습니다.\n\n"
                "설치 방법:\n"
                "1. Ollama 다운로드: https://ollama.com/download\n"
                "2. Python 패키지 설치: pip install ollama\n"
                "3. 모델 다운로드: ollama pull gemma2:2b"
            )
            return
        
        # Summary 필드에서 텍스트 가져오기
        summary_text = self.other_fields['summary'].toPlainText().strip()
        
        if not summary_text:
            QMessageBox.warning(self, "입력 필요", "먼저 Summary 필드에 버그 제목을 입력해주세요.")
            return
        
        # 모델 존재 여부 확인
        try:
            model_exists = AIAssistant.check_model_exists(AIConfig.DEFAULT_MODEL)
        except Exception as e:
            QMessageBox.critical(
                self,
                "Ollama 연결 실패",
                f"Ollama 서비스에 연결할 수 없습니다.\n\n"
                f"오류: {str(e)}\n\n"
                f"해결 방법:\n"
                f"1. Ollama 데스크톱 애플리케이션을 다운로드하세요:\n"
                f"   https://ollama.com/download\n\n"
                f"2. 설치 후 Ollama를 실행하세요.\n\n"
                f"3. 터미널에서 다음 명령어로 모델을 다운로드하세요:\n"
                f"   ollama pull {AIConfig.DEFAULT_MODEL}\n\n"
                f"자세한 내용은 'OLLAMA_설치가이드.md' 파일을 참고하세요."
            )
            return
        
        if not model_exists:
            # 설치된 모델 목록 가져오기
            available_models = AIAssistant.get_available_models()
            
            if not available_models:
                # 모델이 하나도 없는 경우
                QMessageBox.warning(
                    self,
                    "모델 다운로드 필요",
                    f"Ollama가 실행 중이지만 설치된 모델이 없습니다.\n\n"
                    f"터미널에서 다음 명령어로 모델을 다운로드하세요:\n\n"
                    f"ollama pull {AIConfig.DEFAULT_MODEL}\n\n"
                    f"추천 모델:\n"
                    f"• gemma2:2b (가벼움, ~1.6GB)\n"
                    f"• llama3.2:3b (균형, ~2GB)\n"
                    f"• qwen2.5:3b (한국어 좋음, ~2GB)\n\n"
                    f"자세한 내용은 'OLLAMA_설치가이드.md' 파일을 참고하세요."
                )
                return
            else:
                # 다른 모델은 있지만 요청한 모델이 없는 경우
                models_text = "\n".join([f"• {m}" for m in available_models])
                reply = QMessageBox.question(
                    self,
                    "모델 미설치",
                    f"요청한 모델({AIConfig.DEFAULT_MODEL})이 설치되지 않았습니다.\n\n"
                    f"현재 설치된 모델:\n{models_text}\n\n"
                    f"설치하려면 터미널에서 다음 명령어를 실행하세요:\n"
                    f"ollama pull {AIConfig.DEFAULT_MODEL}\n\n"
                    f"그래도 계속하시겠습니까? (첫 번째 설치된 모델 사용)",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )
                
                if reply == QMessageBox.No:
                    return
        
        # 진행 다이얼로그 표시
        self.ai_progress_dialog = QProgressDialog(
            "AI가 버그 세부정보를 생성 중입니다...\n잠시만 기다려주세요.",
            "취소",
            0, 0,  # 무한 진행바
            self
        )
        self.ai_progress_dialog.setWindowTitle("AI 생성 중")
        self.ai_progress_dialog.setWindowModality(Qt.WindowModal)
        self.ai_progress_dialog.setMinimumDuration(0)
        self.ai_progress_dialog.canceled.connect(self._on_ai_generation_canceled)
        self.ai_progress_dialog.show()
        
        # AI 생성 스레드 시작
        self.ai_thread = AIGenerationThread(summary_text, DIR_PRESET)
        self.ai_thread.finished.connect(self._on_ai_generation_finished)
        self.ai_thread.error.connect(self._on_ai_generation_error)
        self.ai_thread.start()
        
        logger.info(f"AI 생성 시작: {summary_text[:50]}...")
    
    def _on_ai_generation_finished(self, result: Dict[str, str]):
        """AI 생성 완료 시 호출"""
        if self.ai_progress_dialog:
            self.ai_progress_dialog.close()
            self.ai_progress_dialog = None
        
        # 생성된 데이터를 필드에 적용
        if 'priority' in result:
            self.widgets['priority'].setCurrentText(result['priority'])
        
        if 'severity' in result:
            self.widgets['severity'].setCurrentText(result['severity'])
        
        if 'steps' in result:
            self.widgets['steps'].setPlainText(result['steps'])
        
        if 'description' in result:
            self.widgets['description'].setPlainText(result['description'])
        
        logger.info("AI 생성 완료 및 필드 적용 완료")
        
        QMessageBox.information(
            self,
            "AI 생성 완료",
            "AI가 버그 세부정보를 성공적으로 생성했습니다.\n"
            "필요한 경우 수동으로 수정해주세요."
        )
    
    def _on_ai_generation_error(self, error_message: str):
        """AI 생성 실패 시 호출"""
        if self.ai_progress_dialog:
            self.ai_progress_dialog.close()
            self.ai_progress_dialog = None
        
        logger.error(f"AI 생성 실패: {error_message}")
        
        QMessageBox.critical(
            self,
            "AI 생성 실패",
            f"AI 생성 중 오류가 발생했습니다:\n\n{error_message}\n\n"
            f"문제 해결:\n"
            f"1. Ollama가 실행 중인지 확인\n"
            f"2. 모델이 다운로드되었는지 확인: ollama list\n"
            f"3. 모델 다운로드: ollama pull {AIConfig.DEFAULT_MODEL}"
        )
    
    def _on_ai_generation_canceled(self):
        """AI 생성 취소 시 호출"""
        if self.ai_thread and self.ai_thread.isRunning():
            self.ai_thread.terminate()
            self.ai_thread.wait()
        
        logger.info("AI 생성 취소됨")
        
        if self.ai_progress_dialog:
            self.ai_progress_dialog.close()
            self.ai_progress_dialog = None
    
    def save_settings(self, filename: str = SETTINGS_FILE):
        """설정을 저장"""
        settings = {
            'sub_label': self.sub_label.text() if hasattr(self, 'sub_label') and self.sub_label else '',
            'priority': self.widgets['priority'].currentText(),
            'severity': self.widgets['severity'].currentText(), 
            'prevalence': self.widgets['prevalence'].currentText(),
            'repro_rate': self.widgets['repro_rate'].currentText(),
            'steps': self.widgets['steps'].toPlainText(),
            'description': self.widgets['description'].toPlainText(),
        }
        
        # 다른 필드들
        for field_name, widget in self.other_fields.items():
            settings[field_name] = widget.text()
        
        self.file_manager.save_json(settings, filename)
    
    def load_settings(self, filename: str = SETTINGS_FILE):
        """설정을 로드"""
        settings = self.file_manager.load_json(filename)
        if settings:
            self._apply_settings_to_widgets(settings)
    
    def _apply_settings_to_widgets(self, settings: dict):
        """설정을 위젯들에 적용"""
        # sub_label 설정
        if hasattr(self, 'sub_label'):
            self.sub_label.setText(settings.get('sub_label', ''))
        
        # 콤보박스들 설정
        self.widgets['priority'].setCurrentText(settings.get('priority', 'Blocker'))
        self.widgets['severity'].setCurrentText(settings.get('severity', '1 - Critical'))
        self.widgets['prevalence'].setCurrentText(settings.get('prevalence', '1 - All users'))
        self.widgets['repro_rate'].setCurrentText(settings.get('repro_rate', '1 - 100% reproducible'))
        
        # 텍스트 필드들 설정
        self.widgets['steps'].setPlainText(settings.get('steps', ''))
        self.widgets['description'].setPlainText(settings.get('description', ''))
        
        # 다른 필드들 설정
        for field_name, widget in self.other_fields.items():
            if hasattr(widget, 'setPlainText'):  # QTextEdit (summary)
                widget.setPlainText(settings.get(field_name, ''))
            elif hasattr(widget, 'setText'):  # QLineEdit
                widget.setText(settings.get(field_name, ''))
            elif hasattr(widget, 'setCurrentText'):  # 콤보박스
                widget.setCurrentText(settings.get(field_name, ''))
    
    def execute(self):
        """JIRA 이슈를 실행"""
        # 현재 설정을 프리셋으로 저장
        self.save_preset()
        
        # 이슈 데이터 준비
        issue_data = self._prepare_issue_data()
        
        # 엑셀 추출 옵션이 활성화되어 있으면 엑셀로 저장
        if self.app_settings.get('excel_export_enabled', True):
            self._export_to_excel(issue_data)
        
        # 스레드에서 실행
        def thread_function():
            create_issue(**issue_data)
        
        issue_thread = threading.Thread(target=thread_function)
        issue_thread.start()
    
    def _prepare_issue_data(self) -> Dict[str, str]:
        """이슈 데이터를 준비"""
        # 라벨 처리
        label = self.other_fields['label'].text()
        sub_label = self.sub_label.text() if hasattr(self, 'sub_label') and self.sub_label else ''
        
        final_label = ""
        if sub_label:
            final_label = f'[{label}][{sub_label}] '
            if hasattr(self, 'include_main_label_check_box') and self.include_main_label_check_box and not self.include_main_label_check_box.isChecked():
                final_label = f'[{sub_label}] '
        else:
            final_label = f'[{label}] '
        
        # summary 필드에서 텍스트 가져오기 (QTextEdit 사용하므로 toPlainText() 사용)
        summary_text = self.other_fields['summary'].toPlainText().replace('\n', ' ').strip()
        summary = f"{final_label}{summary_text}"
        
        def get_field_value(field_name):
            """필드값을 가져오는 헬퍼 함수"""
            widget = self.other_fields.get(field_name)
            if widget:
                if hasattr(widget, 'toPlainText'):  # QTextEdit
                    return widget.toPlainText().replace('\n', ' ').strip()
                elif hasattr(widget, 'text'):  # QLineEdit
                    return widget.text()
                elif hasattr(widget, 'currentText'):  # 콤보박스
                    return widget.currentText()
            return ''
        
        return {
            'summary': summary,
            'linkedIssues': get_field_value('linkedIssues'),
            'issue': get_field_value('issue'),
            'parent': get_field_value('parent'),
            'reviewer': get_field_value('reviewer'),
            'branch': get_field_value('branch'),
            'build': get_field_value('build'),
            'fixversion': get_field_value('fixversion'),
            'component': get_field_value('component'),
            'label': label,
            'priority': self.widgets['priority'].currentText(),
            'severity': self.widgets['severity'].currentText(),
            'prevalence': self.widgets['prevalence'].currentText(),
            'repro_rate': self.widgets['repro_rate'].currentText(),
            'steps': self.widgets['steps'].toPlainText(),
            'description': self.widgets['description'].toPlainText(),
            'team': get_field_value('team')
        }
    
    def refresh_presets(self):
        """프리셋들을 새로고침 (새로운 3단계 구조)"""
        self.widgets['preset_prefix'].clear()
        self.widgets['preset_name'].clear()
        self.widgets['preset_version'].clear()
        
        # 정렬 옵션 확인
        sort_by_date = self.widgets['preset_sort'].currentText() == '최신순'
        
        # 프리셋을 prefix -> name -> versions 구조로 로드
        structure = self.preset_manager.get_preset_names_and_versions()
        
        # Prefix 콤보박스 채우기
        prefixes = list(structure.keys())
        if not sort_by_date:
            prefixes.sort()
        
        self.widgets['preset_prefix'].addItems(prefixes)
        self._structure = structure  # 구조 저장
        
        # 첫 번째 항목 선택
        if prefixes:
            self.widgets['preset_prefix'].setCurrentIndex(0)
            self._on_prefix_changed()
    
    def _on_prefix_changed(self):
        """Prefix가 변경되었을 때"""
        current_prefix = self.widgets['preset_prefix'].currentText()
        if not current_prefix or not hasattr(self, '_structure'):
            return
            
        self.widgets['preset_name'].clear()
        self.widgets['preset_version'].clear()
        
        # 선택된 prefix의 name들을 로드
        if current_prefix in self._structure:
            names = list(self._structure[current_prefix].keys())
            names.sort()
            self.widgets['preset_name'].addItems(names)
            
            # 첫 번째 name 선택
            if names:
                self.widgets['preset_name'].setCurrentIndex(0)
                self._on_name_changed()
    
    def _on_name_changed(self):
        """Name이 변경되었을 때"""
        current_prefix = self.widgets['preset_prefix'].currentText()
        current_name = self.widgets['preset_name'].currentText()
        
        if not current_prefix or not current_name or not hasattr(self, '_structure'):
            return
            
        self.widgets['preset_version'].clear()
        
        # 선택된 name의 버전들을 로드
        if current_prefix in self._structure and current_name in self._structure[current_prefix]:
            versions = self._structure[current_prefix][current_name]
            version_items = []
            
            for version_num, filename in versions:
                if version_num == 0:
                    version_items.append(f"원본 ({filename})")
                else:
                    version_items.append(f"v{version_num} ({filename})")
            
            self.widgets['preset_version'].addItems(version_items)
            
            # 첫 번째 버전(최신) 선택
            if version_items:
                self.widgets['preset_version'].setCurrentIndex(0)
    
    def apply_preset(self):
        """프리셋을 적용"""
        current_prefix = self.widgets['preset_prefix'].currentText()
        current_name = self.widgets['preset_name'].currentText()
        version_text = self.widgets['preset_version'].currentText()
        
        if not current_prefix or not current_name or not version_text:
            return
            
        # 선택된 파일명 찾기
        if hasattr(self, '_structure') and current_prefix in self._structure and current_name in self._structure[current_prefix]:
            versions = self._structure[current_prefix][current_name]
            selected_filename = None
            
            for version_num, filename in versions:
                if version_text.endswith(f"({filename})"):
                    selected_filename = filename
                    break
            
            if selected_filename:
                self.widgets['preset_line'].setText(selected_filename[:-5])  # .json 제거
                settings = self.preset_manager.load_preset(selected_filename)
                if settings:
                    self._apply_settings_to_widgets(settings)
    
    def save_preset(self):
        """프리셋을 저장"""
        new_preset = self.widgets['preset_line'].text().strip()
        if not new_preset:
            return
            
        if not ValidationHelper.is_valid_filename(new_preset):
            QMessageBox.warning(self, "오류", "유효하지 않은 파일명입니다.")
            return
            
        if not new_preset.endswith('.json'):
            new_preset = f'{new_preset}.json'
        
        # 현재 설정 수집
        settings = self._get_current_settings()
        
        # 저장
        if self.preset_manager.save_preset(new_preset, settings):
            logger.info(f'프리셋 저장 성공: {new_preset}')
        else:
            QMessageBox.critical(self, "오류", f"프리셋 '{new_preset}' 저장에 실패했습니다.")
    
    def _get_current_settings(self) -> Dict[str, Any]:
        """현재 설정을 수집"""
        settings = {
            'sub_label': self.sub_label.text() if hasattr(self, 'sub_label') and self.sub_label else '',
            'priority': self.widgets['priority'].currentText(),
            'severity': self.widgets['severity'].currentText(),
            'prevalence': self.widgets['prevalence'].currentText(), 
            'repro_rate': self.widgets['repro_rate'].currentText(),
            'steps': self.widgets['steps'].toPlainText(),
            'description': self.widgets['description'].toPlainText(),
        }
        
        for field_name, widget in self.other_fields.items():
            if hasattr(widget, 'toPlainText'):  # QTextEdit (summary)
                settings[field_name] = widget.toPlainText()
            elif hasattr(widget, 'text'):  # QLineEdit
                settings[field_name] = widget.text()
            elif hasattr(widget, 'currentText'):  # 콤보박스
                settings[field_name] = widget.currentText()
        
        return settings
    
    def delete_preset(self):
        """프리셋을 삭제"""
        current_prefix = self.widgets['preset_prefix'].currentText()
        current_name = self.widgets['preset_name'].currentText()
        version_text = self.widgets['preset_version'].currentText()
        
        if not current_prefix or not current_name or not version_text:
            return

        # 선택된 파일명 찾기
        selected_filename = None
        if hasattr(self, '_structure') and current_prefix in self._structure and current_name in self._structure[current_prefix]:
            versions = self._structure[current_prefix][current_name]
            
            for version_num, filename in versions:
                if version_text.endswith(f"({filename})"):
                    selected_filename = filename
                    break
        
        if not selected_filename:
            return

        # 확인 대화상자
        reply = QMessageBox.question(
            self, "프리셋 삭제", 
            f"프리셋 '{selected_filename}'을(를) 정말 삭제하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            if self.preset_manager.delete_preset(selected_filename):
                self.refresh_presets()
                QMessageBox.information(self, "성공", f"프리셋 '{selected_filename}'이(가) 삭제되었습니다.")
            else:
                QMessageBox.critical(self, "오류", f"프리셋 '{selected_filename}' 삭제에 실패했습니다.")
    
    def show_about_dialog(self):
        """About 대화상자를 표시"""
        about_dialog = QDialog(self)
        about_dialog.setWindowTitle("About")
        layout = QVBoxLayout()
        
        recent_file_name, recent_moditime = FileManager.get_most_recent_file()
        
        info_labels = [
            "Version: v2.0 (Refactored)",
            f"Last update date: {recent_moditime}" if recent_moditime else "Last update date: Unknown",
            "Created by: mssung@pubg.com", 
            "First production date: 2024-07-01",
            "Refactored date: 2024-11-XX"
        ]
        
        for info in info_labels:
            layout.addWidget(QLabel(info, about_dialog))
        
        # 버튼들
        h_layout = QHBoxLayout()
        close_btn = QPushButton("닫기")
        close_btn.clicked.connect(about_dialog.close)
        h_layout.addWidget(close_btn)
        layout.addLayout(h_layout)
        
        about_dialog.setLayout(layout)
        about_dialog.exec_()
    
    def show_settings_dialog(self):
        """Settings 대화상자를 표시"""
        dialog = SettingsDialog(self, self.app_settings)
        if dialog.exec_() == QDialog.Accepted:
            # 설정 저장
            self.app_settings = dialog.get_settings()
            self._save_app_settings()
            logger.info(f"설정 저장됨: {self.app_settings}")
    
    def _load_app_settings(self) -> Dict[str, Any]:
        """앱 설정을 로드"""
        settings = self.file_manager.load_json(APP_SETTINGS_FILE)
        if settings is None:
            # 기본값
            return {'excel_export_enabled': True}
        return settings
    
    def _save_app_settings(self):
        """앱 설정을 저장"""
        self.file_manager.save_json(self.app_settings, APP_SETTINGS_FILE)
    
    def _export_to_excel(self, issue_data: Dict[str, str]):
        """이슈 데이터를 엑셀 파일로 추출"""
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            import os
            
            # result 폴더 확인 및 생성
            os.makedirs(DIR_RESULT, exist_ok=True)
            
            # 엑셀 파일 존재 여부 확인
            if os.path.exists(EXCEL_EXPORT_FILE):
                # 기존 파일 열기
                wb = openpyxl.load_workbook(EXCEL_EXPORT_FILE)
                ws = wb.active
            else:
                # 새 파일 생성
                wb = Workbook()
                ws = wb.active
                ws.title = "Bug Reports"
                
                # 헤더 작성
                headers = [
                    "생성시간", "Summary", "Team", "Linked Issues", "Issue", "Parent",
                    "Reviewer", "Branch", "Build", "Fix Version", "Component",
                    "Label", "Priority", "Severity", "Prevalence", "Repro Rate",
                    "Steps", "Description"
                ]
                ws.append(headers)
                
                # 헤더 스타일 적용
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 현재 시간
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # 데이터 행 추가
            row_data = [
                current_time,
                issue_data.get('summary', ''),
                issue_data.get('team', ''),
                issue_data.get('linkedIssues', ''),
                issue_data.get('issue', ''),
                issue_data.get('parent', ''),
                issue_data.get('reviewer', ''),
                issue_data.get('branch', ''),
                issue_data.get('build', ''),
                issue_data.get('fixversion', ''),
                issue_data.get('component', ''),
                issue_data.get('label', ''),
                issue_data.get('priority', ''),
                issue_data.get('severity', ''),
                issue_data.get('prevalence', ''),
                issue_data.get('repro_rate', ''),
                issue_data.get('steps', ''),
                issue_data.get('description', '')
            ]
            ws.append(row_data)
            
            # 열 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 파일 저장
            wb.save(EXCEL_EXPORT_FILE)
            logger.info(f"버그 정보가 엑셀 파일로 저장되었습니다: {EXCEL_EXPORT_FILE}")
            
        except ImportError:
            logger.error("openpyxl 패키지가 설치되지 않았습니다. pip install openpyxl을 실행하세요.")
            QMessageBox.warning(
                self,
                "패키지 누락",
                "openpyxl 패키지가 필요합니다.\n\n"
                "다음 명령어로 설치하세요:\n"
                "pip install openpyxl"
            )
        except Exception as e:
            logger.error(f"엑셀 파일 저장 중 오류: {e}", exc_info=True)
            QMessageBox.warning(
                self,
                "엑셀 저장 실패",
                f"엑셀 파일 저장 중 오류가 발생했습니다:\n{str(e)}"
            )
    
    def keyPressEvent(self, event):
        """키 이벤트 처리"""
        if event.key() == Qt.Key_F12:
            self.debug_function()
    
    def debug_function(self):
        """디버그 함수 (F12)"""
        logger.info("F12 디버그 함수 실행됨")
        # 필요시 디버그 로직 추가
    
    def open_excel_file(self):
        """엑셀 파일 열기"""
        excel_path = self.excel_widgets['path_input'].text().strip()
        if not excel_path:
            QMessageBox.warning(self, "경로 없음", "엑셀 파일 경로를 입력해주세요.")
            return
        
        import os
        if not os.path.exists(excel_path):
            QMessageBox.warning(self, "파일 없음", f"파일을 찾을 수 없습니다:\n{excel_path}")
            return
        
        try:
            # 엑셀 파일 열기
            os.startfile(excel_path)
            logger.info(f"엑셀 파일 열기: {excel_path}")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"파일 열기 실패:\n{str(e)}")
    
    def execute_excel_batch(self):
        """엑셀 파일의 데이터를 순차적으로 JIRA로 생성 (스레드 기반)"""
        excel_path = self.excel_widgets['path_input'].text().strip()
        if not excel_path:
            QMessageBox.warning(self, "경로 없음", "엑셀 파일 경로를 입력해주세요.")
            return
        
        import os
        if not os.path.exists(excel_path):
            QMessageBox.warning(self, "파일 없음", f"파일을 찾을 수 없습니다:\n{excel_path}")
            return
        
        try:
            import openpyxl
        except ImportError:
            QMessageBox.critical(
                self,
                "패키지 누락",
                "openpyxl 패키지가 필요합니다.\n\n"
                "다음 명령어로 설치하세요:\n"
                "pip install openpyxl"
            )
            return
        
        try:
            # 엑셀 파일 읽기 (데이터 개수 확인용)
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            data_rows = ws.max_row - 1  # 헤더 제외
            
            if data_rows <= 0:
                QMessageBox.warning(self, "데이터 없음", "엑셀 파일에 데이터가 없습니다.")
                return
            
            # 확인 대화상자
            reply = QMessageBox.question(
                self,
                "일괄 실행 확인",
                f"총 {data_rows}개의 이슈를 순차적으로 생성합니다.\n\n"
                f"• 각 이슈는 별도의 브라우저 탭에서 생성됩니다.\n"
                f"• 자동으로 연속 실행되며, 완료 후 각 탭에서 확인/수정 가능합니다.\n"
                f"• 프로그레스바의 '중단' 버튼으로 중지할 수 있습니다.\n\n"
                f"계속하시겠습니까?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
            
            # 진행 다이얼로그 생성
            self.excel_progress_dialog = QProgressDialog(
                "JIRA 이슈 생성 준비 중...",
                "중단",
                0,
                data_rows,
                self
            )
            self.excel_progress_dialog.setWindowTitle("일괄 실행")
            self.excel_progress_dialog.setWindowModality(Qt.WindowModal)
            self.excel_progress_dialog.setMinimumDuration(0)
            self.excel_progress_dialog.canceled.connect(self._on_excel_batch_canceled)
            self.excel_progress_dialog.show()
            
            # 스레드 생성 및 시작
            self.excel_batch_thread = ExcelBatchThread(excel_path)
            self.excel_batch_thread.progress_update.connect(self._on_excel_progress_update)
            self.excel_batch_thread.issue_created.connect(self._on_excel_issue_created)
            self.excel_batch_thread.error_occurred.connect(self._on_excel_error_occurred)
            self.excel_batch_thread.finished.connect(self._on_excel_batch_finished)
            self.excel_batch_thread.start()
            
            logger.info(f"엑셀 일괄 실행 시작: {excel_path} ({data_rows}개)")
            
        except Exception as e:
            logger.error(f"엑셀 일괄 실행 준비 실패: {e}", exc_info=True)
            QMessageBox.critical(
                self,
                "오류",
                f"엑셀 파일 처리 중 오류가 발생했습니다:\n\n{str(e)}"
            )
    
    def _on_excel_progress_update(self, value: int, message: str):
        """진행 상황 업데이트"""
        if self.excel_progress_dialog:
            self.excel_progress_dialog.setValue(value)
            self.excel_progress_dialog.setLabelText(message)
    
    def _on_excel_issue_created(self, current: int, total: int, summary: str):
        """이슈 생성 완료 시 호출"""
        logger.info(f"이슈 생성 완료 ({current}/{total}): {summary[:50]}")
    
    def _on_excel_error_occurred(self, row_num: int, error_message: str):
        """에러 발생 시 호출"""
        logger.error(f"이슈 생성 실패 (행 {row_num}): {error_message}")
        
        # 에러 발생 시 사용자에게 물어봄
        reply = QMessageBox.question(
            self,
            "오류 발생",
            f"이슈 생성 중 오류가 발생했습니다 (행 {row_num}):\n\n{error_message[:200]}\n\n"
            f"계속하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        
        if reply != QMessageBox.Yes:
            # 사용자가 중단을 선택한 경우
            if self.excel_batch_thread:
                self.excel_batch_thread.cancel()
    
    def _on_excel_batch_finished(self, success_count: int, failed_rows: list):
        """일괄 실행 완료 시 호출"""
        # 프로그레스바 닫기
        if self.excel_progress_dialog:
            self.excel_progress_dialog.close()
            self.excel_progress_dialog = None
        
        # 결과 메시지
        result_message = f"일괄 실행 완료!\n\n"
        result_message += f"✅ 성공: {success_count}개\n"
        
        if failed_rows:
            result_message += f"❌ 실패: {len(failed_rows)}개\n\n"
            result_message += "실패한 행:\n"
            for row_num, error in failed_rows[:5]:  # 최대 5개만 표시
                result_message += f"- 행 {row_num}: {error[:50]}...\n"
            if len(failed_rows) > 5:
                result_message += f"... 외 {len(failed_rows) - 5}개\n"
        
        result_message += f"\n💡 브라우저의 각 탭에서 생성된 이슈를 확인하고\n"
        result_message += f"필요시 수정한 후 'Create' 버튼을 눌러주세요."
        
        QMessageBox.information(self, "일괄 실행 완료", result_message)
        
        # 스레드 정리
        self.excel_batch_thread = None
    
    def _on_excel_batch_canceled(self):
        """일괄 실행 취소 시 호출"""
        if self.excel_batch_thread and self.excel_batch_thread.isRunning():
            self.excel_batch_thread.cancel()
            # 스레드가 종료될 때까지 대기 (최대 3초)
            self.excel_batch_thread.wait(3000)
        
        logger.info("엑셀 일괄 실행 취소됨")
        
        if self.excel_progress_dialog:
            self.excel_progress_dialog.close()
            self.excel_progress_dialog = None
    
    def closeEvent(self, event):
        """창 닫기 이벤트 처리"""
        self.save_settings()
        event.accept()


def main():
    """메인 함수"""
    app = QApplication(sys.argv)
    
    # 애플리케이션 설정
    app.setApplicationName('JIRA Bug Report Tool')
    app.setApplicationVersion('2.0')
    app.setOrganizationName('PUBG Corporation')
    
    # 메인 윈도우 생성
    window = BugReportApp()
    
    # 이벤트 루프 시작
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
